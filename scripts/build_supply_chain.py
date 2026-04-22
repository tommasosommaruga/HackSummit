"""
Build the unified REE supply-chain dataset.

Inputs  (in data/raw/):
  - Global_REE_combined.xlsx              — USGS occurrences (deposits).
  - Company that processes REE.xlsx       — "REE projects" sheet (commercial
                                            projects, 147 rows, no lat/lon)
                                            + "Factory" sheet (67 plants with
                                            Upstream/Downstream text refs).

Outputs (in data/processed/ and mirrored to webapp/public/):
  - nodes.json            — unified nodes: deposits + projects + refineries.
  - edges.json            — deposit→project (from fuzzy join) and
                            project→refinery / refinery→refinery / refinery→OEM
                            (from Factory sheet Upstream/Downstream).
  - supply_chain.json     — {nodes, edges, meta} bundle, single-fetch friendly.
  - geocode_cache.json    — persistent Nominatim cache (shared across runs).

Schema (also documented in README.md):

  NODE:
    id: str           (globally unique: dep_<ID_No>, proj_<n>, fac_<n>, …)
    type: str         ('deposit' | 'project' | 'refinery' | 'oem' | 'reseller')
    name: str
    lat, lng: float
    country: str
    state: str?
    region: str?
    geocoded: bool    (true = placed from explicit coords OR confident lookup)
    precision: str    ('exact' | 'state' | 'country' | 'joined')
    company: str?
    status: str?      (canonical textual status — source-dependent)
    status_code: str? (numeric/alpha code from File 2 / Factory sheet)
    deposit_type: str?
    commodities: str?
    ree_grade: str?
    resource_kt_reo: float?
    grade_pct: float?
    capacity: str?    (refineries — free-text from Factory sheet)
    yield: str?
    ref_urls: list[str]
    joined_to: str?   (node id this row joined to — deposit↔project linkage)

  EDGE:
    id: str
    from_id: str
    to_id: str
    material: str?    ('ore' | 'concentrate' | 'mixed_reo' | 'separated_reo' | …)
    volume_tons_per_year: float?
    year: int?
    source: str       ('fuzzy_join' | 'factory_upstream' | 'factory_downstream')
    evidence: str?    (original upstream/downstream text for debugging)
"""
from __future__ import annotations
import json
import math
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OCC_XLSX = ROOT / 'data' / 'raw' / 'Global_REE_combined.xlsx'
CO_XLSX  = ROOT / 'data' / 'raw' / 'Company that processes REE.xlsx'
CN_XLSX  = ROOT / 'data' / 'raw' / 'China_REE_Refineries.xlsx'
PROC     = ROOT / 'data' / 'processed'
WEB      = ROOT / 'webapp' / 'public'
GEOCACHE = PROC / 'geocode_cache.json'


# ── Nominatim geocoder (cached, rate-limited) ─────────────────────────────────
_LAST_REQ = [0.0]
_UA = 'HackSummit-REE-Geocoder/1.0 (https://github.com/mmsellam)'


def _load_cache():
    if GEOCACHE.exists():
        return json.loads(GEOCACHE.read_text())
    return {}


def _save_cache(cache):
    GEOCACHE.parent.mkdir(parents=True, exist_ok=True)
    GEOCACHE.write_text(json.dumps(cache, indent=2, ensure_ascii=False, sort_keys=True))


_QUERIES = [0]
_HITS = [0]


def nominatim_lookup(query, cache):
    if not query:
        return None
    if query in cache:
        return cache[query]
    wait = 1.05 - (time.monotonic() - _LAST_REQ[0])
    if wait > 0:
        time.sleep(wait)
    url = ('https://nominatim.openstreetmap.org/search?'
           + urllib.parse.urlencode({'q': query, 'format': 'json', 'limit': 1}))
    req = urllib.request.Request(url, headers={'User-Agent': _UA})
    _QUERIES[0] += 1
    print(f'  [geo #{_QUERIES[0]:>4}] {query[:70]}', flush=True)
    try:
        with urllib.request.urlopen(req, timeout=8) as resp:
            results = json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        print(f'         error: {e}', flush=True)
        cache[query] = None
        _LAST_REQ[0] = time.monotonic()
        # Flush cache periodically so progress survives crashes.
        if _QUERIES[0] % 20 == 0:
            _save_cache(cache)
        return None
    _LAST_REQ[0] = time.monotonic()
    if not results:
        cache[query] = None
        if _QUERIES[0] % 20 == 0:
            _save_cache(cache)
        return None
    try:
        coord = [float(results[0]['lat']), float(results[0]['lon'])]
    except (KeyError, ValueError):
        cache[query] = None
        return None
    cache[query] = coord
    _HITS[0] += 1
    if _QUERIES[0] % 20 == 0:
        _save_cache(cache)
    return coord


# ── Country / state centroids (offline fallback) ──────────────────────────────
# Only used when Nominatim fails. Kept short; Nominatim handles the rest.
COUNTRY_CENTROIDS = {
    'Algeria': (28.03, 1.66), 'Argentina': (-38.42, -63.62), 'Armenia': (40.07, 45.04),
    'Australia': (-25.27, 133.78), 'Bangladesh': (23.68, 90.36), 'Belarus': (53.71, 27.95),
    'Benin': (9.31, 2.32), 'Bolivia': (-16.29, -63.59), 'Brazil': (-14.24, -51.93),
    'Burundi': (-3.37, 29.92), 'Canada': (56.13, -106.35),
    'Central African Republic': (6.61, 20.94), 'Chile': (-35.68, -71.54),
    'China': (35.86, 104.20), 'Colombia': (4.57, -74.30),
    'Czech Republic': (49.82, 15.47),
    'Democratic Republic of Congo': (-4.04, 21.76),
    'Democratic Republic of the Congo': (-4.04, 21.76),
    'Zaire': (-4.04, 21.76), 'Egypt': (26.82, 30.80), 'Ethiopia': (9.14, 40.49),
    'Finland': (61.92, 25.75), 'France': (46.23, 2.21), 'Gabon': (-0.80, 11.61),
    'Germany': (51.17, 10.45), 'Greece': (39.07, 21.82), 'Greenland': (71.71, -42.60),
    'Guyana': (4.86, -58.93), 'India': (20.59, 78.96), 'Indonesia': (-0.79, 113.92),
    'Iran': (32.43, 53.69), 'Italy': (41.87, 12.57), 'Jamaica': (18.11, -77.30),
    'Japan': (36.20, 138.25), 'Kazakhstan': (48.02, 66.92), 'Kenya': (-0.02, 37.91),
    'Kyrgyzstan': (41.20, 74.77), 'Latvia': (56.88, 24.60), 'Madagascar': (-18.77, 46.87),
    'Malawi': (-13.25, 34.30), 'Malaysia': (4.21, 101.98), 'Mexico': (23.63, -102.55),
    'Mongolia': (46.86, 103.85), 'Montenegro': (42.71, 19.37), 'Mozambique': (-18.67, 35.53),
    'Myanmar': (21.91, 95.96), 'Namibia': (-22.96, 18.49), 'Nepal': (28.39, 84.12),
    'New Zealand': (-40.90, 174.89), 'Nigeria': (9.08, 8.68), 'Norway': (60.47, 8.47),
    'Paraguay': (-23.44, -58.44), 'Peru': (-9.19, -75.02), 'Romania': (45.94, 24.97),
    'Russian Federation': (61.52, 105.32), 'Russia': (61.52, 105.32),
    'Rwanda': (-1.94, 29.87), 'Saudi Arabia': (23.89, 45.08), 'Sierra Leone': (8.46, -11.78),
    'South Africa': (-30.56, 22.94), 'South Korea': (35.91, 127.77), 'Spain': (40.46, -3.75),
    'Sri Lanka': (7.87, 80.77), 'Sudan': (12.86, 30.22), 'Swaziland': (-26.52, 31.47),
    'Sweden': (60.13, 18.64), 'Taiwan': (23.70, 120.96), 'Tajikistan': (38.86, 71.28),
    'Tanzania': (-6.37, 34.89), 'Thailand': (15.87, 100.99), 'Tibet': (31.69, 88.09),
    'Tunisia': (33.89, 9.54), 'Turkey': (38.96, 35.24), 'Uganda': (1.37, 32.29),
    'Ukraine': (48.38, 31.17), 'United Arab Emirates': (23.42, 53.85),
    'United Kingdom': (55.38, -3.44), 'UK': (55.38, -3.44),
    'United States': (37.09, -95.71), 'USA': (37.09, -95.71),
    'Uzbekistan': (41.38, 64.59), 'Venezuela': (6.42, -66.59), 'Vietnam': (14.06, 108.28),
    'Zambia': (-13.13, 27.85), 'Zimbabwe': (-19.02, 29.15),
    'Estonia': (58.60, 25.01), 'Belgium': (50.50, 4.47), 'Poland': (51.92, 19.13),
    'Denmark': (56.26, 9.50), 'Netherlands': (52.13, 5.29), 'Ireland': (53.14, -7.69),
    'Hungary': (47.16, 19.50), 'Austria': (47.52, 14.55),
}


def _jitter(base, level, seed):
    """Deterministic jitter keyed by seed so points don't stack."""
    jitter = 0.25 if level == 'state' else (0.6 if level == 'joined' else 1.2)
    n = 0
    for ch in str(seed):
        n = (n * 131 + ord(ch)) & 0xffffff
    angle = (n * 137.508) * math.pi / 180.0
    r = jitter * math.sqrt(((n >> 8) & 0xff) / 255.0)
    return (base[0] + r * math.sin(angle), base[1] + r * math.cos(angle))


def resolve_coord(query_strings, country, state, seed, cache):
    """Try Nominatim for each query in order; fall back to country centroid."""
    for q in query_strings:
        if not q:
            continue
        hit = nominatim_lookup(q, cache)
        if hit:
            return (hit[0], hit[1], 'exact' if ',' in q else 'state')
    if state and country:
        hit = nominatim_lookup(f'{state}, {country}', cache)
        if hit:
            lat, lng = _jitter((hit[0], hit[1]), 'state', seed)
            return (lat, lng, 'state')
    if country and country in COUNTRY_CENTROIDS:
        lat, lng = _jitter(COUNTRY_CENTROIDS[country], 'country', seed)
        return (lat, lng, 'country')
    return None


# ── Cleaners ─────────────────────────────────────────────────────────────────
def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s not in ('nan', 'NaN', 'None', '-') else None


def clean_status_code(v):
    """Normalize raw status cell → canonical short code.

    Examples:
      '4.0'                             → '4'
      '4 - Active production'           → '4'
      '5 - Metal/alloy production (…)'  → '5'
      'T,3(2023)'                       → '3'   (most recent wins)
      'P, 4(?)'                         → '4'
      'Showing(?)'                      → 'Showing'
    """
    s = clean(v)
    if s is None:
        return None
    # '4 - Active production' → '4'  (strip the explanatory tail)
    s = re.split(r'\s+-\s+', s, maxsplit=1)[0].strip()
    # '4.0' → '4'
    if re.fullmatch(r'-?\d+\.0+', s):
        return s.split('.')[0]
    # Compound like 'T,3(2023)' → take the rightmost token (most recent).
    tokens = [t.strip() for t in re.split(r'\s*[,&]\s*', s) if t.strip()]
    if tokens:
        s = tokens[-1]
    # Strip year-paren suffix '3(2023)' → '3'.
    s = re.sub(r'\s*\(.*?\)\s*$', '', s)
    # Drop trailing '.0' from compound floats.
    s = re.sub(r'\.0+$', '', s)
    # USGS '(?)' uncertainty marker.
    s = re.sub(r'\s*\(\?\)\s*$', '', s)
    return s or None


def num(v):
    if v is None:
        return None
    try:
        f = float(v)
        return f if math.isfinite(f) else None
    except (TypeError, ValueError):
        return None


def normalize_name(s):
    """Lowercase, strip non-alphanumeric, collapse whitespace. That's it.
    No token substitutions, no word drops — those are fuzzy heuristics that
    produce false matches."""
    if not s:
        return ''
    s = s.lower()
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return ' '.join(s.split()).strip()


def names_equal(a, b):
    """Exact equality after normalization. No similarity scores."""
    return normalize_name(a) == normalize_name(b) if a and b else False


# ── Sheet loaders ────────────────────────────────────────────────────────────
def load_occurrences(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['All Occurrences']
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    idx = {h: i for i, h in enumerate(headers)}
    out = []
    for r in rows:
        out.append({
            'id_no':    clean(r[idx['ID_No']]),
            'name':     clean(r[idx['Name']]),
            'country':  clean(r[idx['Country']]),
            'state':    clean(r[idx['State_Prov']]),
            'lat':      num(r[idx['Latitude']]),
            'lng':      num(r[idx['Longitude']]),
            'dep_type': clean(r[idx['Dep_Type']]),
            'status':   clean(r[idx['Status']]),
            'p_status': clean(r[idx['P_Status']]),
            'company':  clean(r[idx['Company']]),
            'commods':  clean(r[idx['Commods']]),
            'ree':      clean(r[idx['REE']]),
            'region':   clean(r[idx['Region']]),
        })
    return out


def load_projects(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['REE projects']
    rows = list(ws.iter_rows(values_only=True))
    # rows[0] is the supplementary-table title; rows[1] are the real headers.
    headers = rows[1]
    idx = {h: i for i, h in enumerate(headers)}
    status_col = next(h for h in headers if h and 'Status (2022)' in str(h))
    resource_col = next(h for h in headers if h and 'Resource' in str(h))
    grade_col = next(h for h in headers if h and h.strip() == 'Grade (wt. %)')
    out = []
    for r in rows[2:]:
        pno = clean(r[idx['Project No.']])
        if not pno:
            continue
        out.append({
            'pno':          pno,
            'company':      clean(r[idx['Company Name']]),
            'name':         clean(r[idx['Project Name']]),
            'location':     clean(r[idx['Location']]),
            'continent':    clean(r[idx['Continent']]),
            'status_code':  clean_status_code(r[idx[status_col]]),
            'dep_type':     clean(r[idx['Deposit type']]),
            'resource':     num(r[idx[resource_col]]),
            'grade':        num(r[idx[grade_col]]),
            'ref_project':  clean(r[idx['Project and status Ref.']])
                            if 'Project and status Ref.' in idx else None,
        })
    return out


def parse_quota(s):
    """Pull the first numeric tonnage out of a free-text quota string.
    '170,001' → 170001;  '83,999 (full year, incl. X)' → 83999;
    'Not separately published' / 'Absorbed into …' → None.
    """
    if not s:
        return None
    m = re.search(r'[\d][\d,]*(?:\.\d+)?', str(s))
    if not m:
        return None
    try:
        return float(m.group(0).replace(',', ''))
    except ValueError:
        return None


def load_china_refineries(path):
    """Load the China_REE_Refineries.xlsx supplement. Returns one dict per
    non-empty row of the 'China Refineries' sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['China Refineries']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers)}

    def col(h): return idx[h] if h in idx else None

    out = []
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        out.append({
            'no':            clean(r[col('No.')]),
            'company':       clean(r[col('Company')]),
            'chinese_name':  clean(r[col('Chinese Name')]),
            'name':          clean(r[col('Project / Plant')]),
            'location':      clean(r[col('Location')]),
            'ticker':        clean(r[col('Stock Ticker')]),
            'ownership':     clean(r[col('Ownership')]),
            'status_text':   clean(r[col('Status (2024)')]),
            'ree_type':      clean(r[col('REE Type')]),
            'smelt_quota':   parse_quota(r[col('Smelting Quota 2024 (t REO)')]),
            'mine_quota':    parse_quota(r[col('Mining Quota 2024 (t REO)')]),
            'capacity':      clean(r[col('Capacity (declared)')]),
            'revenue':       clean(r[col('Revenue 2024 (CNY)')]),
            'net_profit':    clean(r[col('Net Profit 2024 (CNY)')]),
            'op_margin':     clean(r[col('Operating Margin 2024)')]) if 'Operating Margin 2024)' in idx else clean(r[col('Operating Margin 2024')]) if 'Operating Margin 2024' in idx else None,
            'employees':     clean(r[col('Employees')]),
            'upstream':      clean(r[col('Upstream (Feed)')]),
            'downstream':    clean(r[col('Downstream (Clients)')]),
            'products':      clean(r[col('Products')]),
            'disclosure':    clean(r[col('Disclosure Level')]),
            'ref_status':    clean(r[col('Status Ref.')]),
            'ref_capacity':  clean(r[col('Capacity Ref.')]),
            'ref_upstream':  clean(r[col('Upstream Ref.')]),
            'ref_downstream':clean(r[col('Downstream Ref.')]),
            'notes':         clean(r[col('Notes')]),
        })
    return out


def load_factories(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Factory']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    idx = {h: i for i, h in enumerate(headers)}
    status_col = next(h for h in headers if h and 'Status (2022)' in str(h))
    out = []
    last_company = None
    for r in rows[1:]:
        no = clean(r[idx['No.']])
        if not no:
            continue
        co = clean(r[idx['Company']])
        if co:
            last_company = co
        else:
            co = last_company
        out.append({
            'no':          no,
            'company':     co,
            'name':        clean(r[idx['Project']]),
            'location':    clean(r[idx['Location']]),
            'status_code': clean_status_code(r[idx[status_col]]),
            'capacity':    clean(r[idx['Capacity']]),
            'yield':       clean(r[idx['Yield']]),
            'upstream':    clean(r[idx['Upstream']]),
            'downstream':  clean(r[idx['Downstream']]),
            'ref_status':  clean(r[idx['Status Ref.']]),
            'ref_capacity':clean(r[idx['Capacity Ref.']]),
        })
    return out


# ── Build unified nodes ──────────────────────────────────────────────────────
def deposit_country(country):
    """Normalize odd country strings (USGS has 'Russian Federation(?)' etc.)."""
    if not country:
        return None
    return country.rstrip(' (?)').strip()


# Canonical country names we'll look for inside any location string.
# Multi-word entries win over single-word entries (checked first).
_COUNTRY_GAZETTEER = [
    ('United States', ['united states', 'united states of america', 'usa', 'u.s.a', 'u.s.']),
    # 'england'/'scotland'/'wales' omitted — 'Wales' collides with 'New South
    # Wales' (Australia). 'uk' and 'britain' are safe with word-boundary match
    # (won't fire inside 'Ukraine' / 'Britannica').
    ('United Kingdom', ['united kingdom', 'uk', 'great britain', 'britain']),
    ('South Korea',    ['south korea', 'republic of korea', 'korea, rep']),
    ('North Korea',    ['north korea']),
    ('South Africa',   ['south africa']),
    ('Czech Republic', ['czech republic', 'czechia']),
    ('Russian Federation', ['russian federation', 'russia']),
    ('New Zealand',    ['new zealand']),
    ('Saudi Arabia',   ['saudi arabia']),
    ('United Arab Emirates', ['united arab emirates', 'uae']),
    ('Sri Lanka',      ['sri lanka']),
    ('Central African Republic', ['central african republic']),
    ('Sierra Leone',   ['sierra leone']),
    ('Democratic Republic of the Congo',
         ['democratic republic of the congo', 'democratic republic of congo', 'drc', 'zaire']),
    ('Saint Helena, Ascension, and Tristan da Cunha',
         ['saint helena', 'st helena']),
]
# Single-word countries (derived from COUNTRY_CENTROIDS minus the multi-word
# ones above) — matched with word boundaries so "Korea" in "Korea, South" hits
# but "Iran" inside a random word does not.
_SINGLE_COUNTRIES = sorted(
    {c for c in COUNTRY_CENTROIDS if ' ' not in c and len(c) >= 4},
    key=len, reverse=True,
)


def country_from_location(loc):
    """Extract a canonical country name from free-text location.

    All alias checks use word-boundary regex so that 'uk' does not match
    inside 'Ukraine', 'iran' does not match 'Iranian', etc.
    """
    if not loc:
        return None
    text = str(loc).lower()
    # Multi-word / synonym matches first (longer phrases tried first).
    for canonical, aliases in _COUNTRY_GAZETTEER:
        for a in aliases:
            if re.search(rf'\b{re.escape(a)}\b', text):
                return canonical
    # Single-word country match.
    for c in _SINGLE_COUNTRIES:
        if re.search(rf'\b{re.escape(c.lower())}\b', text):
            return c
    # Legacy: last comma-separated token, may not be a real country.
    tail = str(loc).split(',')[-1].strip()
    return tail or None


# Manual coordinate overrides for the long tail of mis-geocoded plants whose
# location strings are too idiosyncratic for the gazetteer (e.g. "Kuantan" with
# no country token at all). Keyed by the produced node id. Use sparingly.
LOCATION_OVERRIDES = {
    'fac_2.0': (3.9745, 103.2869, 'Kuantan, Pahang, Malaysia'),    # Lynas Kuantan Plant
}


def build_deposit_nodes(occurrences, cache):
    nodes = []
    for o in occurrences:
        seed = o['id_no'] or o['name'] or ''
        if o['lat'] is not None and o['lng'] is not None:
            lat, lng, precision = o['lat'], o['lng'], 'exact'
        else:
            # Deposit names are usually mine / site names — not place names — so
            # Nominatim won't resolve them. Skip the name query and go straight
            # to state → country fallback. That's the 820-row path; cuts ~2000
            # Nominatim requests off the end of each run.
            resolved = resolve_coord(
                [], deposit_country(o['country']), o['state'], seed, cache,
            )
            if not resolved:
                continue
            lat, lng, precision = resolved
        nodes.append({
            'id':           f'dep_{o["id_no"]}',
            'type':         'deposit',
            'name':         o['name'] or f'Occurrence #{o["id_no"]}',
            'lat':          round(lat, 5),
            'lng':          round(lng, 5),
            'country':      deposit_country(o['country']),
            'state':        o['state'],
            'region':       o['region'],
            'precision':    precision,
            'geocoded':     True,
            'company':      o['company'],
            'status':       o['status'],
            'pstatus':      o['p_status'],      # USGS P_Status — drives 'active' filter
            'deposit_type': o['dep_type'],
            'commodities':  o['commods'],
            'ree_grade':    o['ree'],
            'ref_urls':     [],
        })
    return nodes


def build_project_nodes(projects, deposit_nodes, cache):
    """Projects get coords from (a) EXACT name+country match to a deposit, or
    (b) Nominatim on the free-text location. No fuzzy matching is used; a
    join is emitted only when the project name equals a deposit name (after
    case/punctuation normalization) and both sit in the same country."""
    # Exact lookup index: (normalized_name, country) → deposit
    deposit_index = {}
    for d in deposit_nodes:
        key = (normalize_name(d['name']), d.get('country'))
        deposit_index.setdefault(key, d)   # first wins if duplicates exist

    nodes = []
    edges = []
    for p in projects:
        country = country_from_location(p['location']) or p['continent']
        if country and country.lower() in ('usa', 'u.s.a.', 'united states of america'):
            country = 'United States'
        if country and country.lower() in ('uk', 'great britain'):
            country = 'United Kingdom'

        # (1) Exact join to a deposit.
        best = deposit_index.get((normalize_name(p['name']), country))
        joined_id = None
        lat = lng = precision = None
        if best is not None:
            joined_id = best['id']
            lat, lng = _jitter((best['lat'], best['lng']), 'joined', p['pno'])
            precision = 'joined'

        # (2) Nominatim fallback on the free-text location.
        if lat is None:
            resolved = resolve_coord(
                [p['location'], f'{p["name"]}, {country}' if p['name'] and country else None],
                country, None, p['pno'], cache,
            )
            if resolved:
                lat, lng, precision = resolved

        if lat is None:
            # keep the record but flag it; will not be rendered
            nodes.append({
                'id':           f'proj_{p["pno"]}',
                'type':         'project',
                'name':         p['name'],
                'lat':          None,
                'lng':          None,
                'country':      country,
                'geocoded':     False,
                'precision':    'unknown',
                'company':      p['company'],
                'status_code':  p['status_code'],
                'deposit_type': p['dep_type'],
                'resource_kt_reo': p['resource'],
                'grade_pct':    p['grade'],
                'location_text': p['location'],
                'ref_urls':     [p['ref_project']] if p.get('ref_project') else [],
                'joined_to':    None,
            })
            continue

        proj = {
            'id':           f'proj_{p["pno"]}',
            'type':         'project',
            'name':         p['name'],
            'lat':          round(lat, 5),
            'lng':          round(lng, 5),
            'country':      country,
            'region':       p['continent'],
            'geocoded':     True,
            'precision':    precision,
            'company':      p['company'],
            'status_code':  p['status_code'],
            'deposit_type': p['dep_type'],
            'resource_kt_reo': p['resource'],
            'grade_pct':    p['grade'],
            'location_text': p['location'],
            'ref_urls':     [p['ref_project']] if p.get('ref_project') else [],
            'joined_to':    joined_id,
        }
        nodes.append(proj)

        if joined_id:
            edges.append({
                'id':       f'edge_join_{p["pno"]}',
                'from_id':  joined_id,
                'to_id':    proj['id'],
                'material': 'ore',
                'volume_tons_per_year': None,
                'year':     2022,
                'source':   'name_join',
                'evidence': f'exact name+country match: {p["name"]}',
            })
    return nodes, edges


def build_factory_nodes(factories, existing_nodes, cache):
    """Factory rows → refinery nodes.

    Coord resolution order:
      1. Upstream text fuzzy-matches an existing node (deposit/project) → adopt
         that node's coords with small jitter. (Mount Weld Plant → Mount Weld
         mine, which Nominatim otherwise resolves to a Tasmania mountain.)
      2. Nominatim on the free-text location string.
      3. Nominatim on "<plant name>, <country>".
    """
    # Exact name → node index (normalized form).
    name_idx = {}
    for n in existing_nodes:
        if n.get('name') and n.get('lat') is not None:
            name_idx.setdefault(normalize_name(n['name']), n)

    nodes = []
    for f in factories:
        location = f['location']
        factory_country = country_from_location(location)
        seed = f['no']
        lat = lng = precision = None

        # (1) Co-locate with upstream ONLY when both endpoints are in the same
        # country. Upstream text like 'Dubbo' (Australia) feeding the Korean
        # Metals Plant (South Korea) must NOT drag the plant to Australia.
        up = f.get('upstream')
        if up:
            primary = re.split(r',| and ', up)[0].strip()
            hit = name_idx.get(normalize_name(primary)) if primary else None
            if hit and hit.get('country') and factory_country and hit['country'] == factory_country:
                lat, lng = _jitter((hit['lat'], hit['lng']), 'joined', seed)
                precision = 'joined'

        # (2) Fall back to Nominatim on location / name.
        if lat is None:
            resolved = resolve_coord(
                [location, f'{f["name"]}, {country_from_location(location)}' if f['name'] else None],
                country_from_location(location), None, seed, cache,
            )
            if resolved:
                lat, lng, precision = resolved

        if lat is None:
            nodes.append({
                'id':       f'fac_{f["no"]}',
                'type':     'refinery',
                'name':     f['name'],
                'lat': None, 'lng': None,
                'country':  country_from_location(location),
                'geocoded': False,
                'precision':'unknown',
                'company':  f['company'],
                'status_code': f['status_code'],
                'capacity': f['capacity'],
                'yield':    f['yield'],
                'location_text': location,
                '_upstream_text':   f['upstream'],
                '_downstream_text': f['downstream'],
                'ref_urls': [u for u in (f['ref_status'], f['ref_capacity']) if u],
            })
            continue
        nodes.append({
            'id':       f'fac_{f["no"]}',
            'type':     'refinery',
            'name':     f['name'],
            'lat':      round(lat, 5),
            'lng':      round(lng, 5),
            'country':  country_from_location(location),
            'geocoded': True,
            'precision':precision,
            'company':  f['company'],
            'status_code': f['status_code'],
            'capacity': f['capacity'],
            'yield':    f['yield'],
            'location_text': location,
            '_upstream_text':   f['upstream'],
            '_downstream_text': f['downstream'],
            'ref_urls': [u for u in (f['ref_status'], f['ref_capacity']) if u],
        })
    return nodes


# ── China refineries (supplemental dataset, 2024 data) ──────────────────────
# A small subset of the 10 rows aren't strictly refineries. The classification
# table below overrides the default 'refinery' type so Bayan Obo (a mine) and
# the NdFeB alloy plant (downstream magnet maker) land in the right bucket.
CHINA_ROW_OVERRIDES = {
    '8':  {'type': 'project',       'note': 'Mine + concentrate producer (RE as byproduct of iron ore). Reclassified from refinery.'},
    '10': {'type': 'magnet_maker',  'note': 'NdFeB alloy / magnet producer. Reclassified from refinery.'},
}

# HQ-only / trading-only rows — no physical refinery to place. Omitted entirely.
CHINA_EXCLUDE = {'7'}

# Multi-site aggregates. Each entry splits the raw row into N refinery nodes,
# one per operational site (HQ and trading arms are intentionally excluded).
# Children share the row's company / quota / revenue metadata and carry a
# shared `group_id` so edges referencing the aggregate fan out to all sites.
CHINA_SITE_SPLITS = {
    # Southern HREE Smelting & Separation Network — 8 provinces, HQ Ganzhou
    '2': [
        {'label': 'Jiangxi',   'geo_query': 'Ganzhou, Jiangxi, China'},   # primary
        {'label': 'Guangxi',   'geo_query': 'Nanning, Guangxi, China'},
        {'label': 'Hunan',     'geo_query': 'Changsha, Hunan, China'},
        {'label': 'Sichuan',   'geo_query': 'Leshan, Sichuan, China'},
        {'label': 'Guangdong', 'geo_query': 'Heyuan, Guangdong, China'},
        {'label': 'Fujian',    'geo_query': 'Longyan, Fujian, China'},
        {'label': 'Yunnan',    'geo_query': 'Kunming, Yunnan, China'},
        {'label': 'Shandong',  'geo_query': 'Jinan, Shandong, China'},
    ],
    # Shenghe Resources — Leshan (primary smelting) + Jiangxi (HREE). Singapore
    # is explicitly a trading hub, so it is NOT split into its own site.
    '3': [
        {'label': 'Leshan, Sichuan', 'geo_query': 'Leshan, Sichuan, China'},
        {'label': 'Jiangxi',         'geo_query': 'Ganzhou, Jiangxi, China'},
    ],
    # Xiamen Tungsten — three Fujian cities
    '4': [
        {'label': 'Xiamen, Fujian',  'geo_query': 'Xiamen, Fujian, China'},
        {'label': 'Longyan, Fujian', 'geo_query': 'Longyan, Fujian, China'},
        {'label': 'Sanming, Fujian', 'geo_query': 'Sanming, Fujian, China'},
    ],
}


def _base_china_node(r, node_id, name_override=None):
    """Build the shared metadata shell for a China refinery node."""
    override = CHINA_ROW_OVERRIDES.get(r['no'], {})
    node_type = override.get('type', 'refinery')
    node = {
        'id':           node_id,
        'type':         node_type,
        'name':         name_override or r['name'],
        'lat':          None,
        'lng':          None,
        'country':      'China',
        'geocoded':     False,
        'precision':    'unknown',
        'confidence':   'inferred',
        'company':      r['company'],
        'chinese_name': r['chinese_name'],
        'ticker':       r['ticker'],
        'ownership':    r['ownership'],
        'status_code':  clean_status_code(r['status_text']) if r['status_text'] else None,
        'status_text':  r['status_text'],
        'ree_type':     r['ree_type'],
        'smelting_quota_t_reo_2024': r['smelt_quota'],
        'mining_quota_t_reo_2024':   r['mine_quota'],
        'capacity':     r['capacity'],
        'revenue_2024_cny':    r['revenue'],
        'net_profit_2024_cny': r['net_profit'],
        'operating_margin_2024': r['op_margin'],
        'employees':    r['employees'],
        'products':     r['products'],
        'disclosure_level': r['disclosure'],
        'location_text': r['location'],
        'data_year':    2024,
        'source_file':  'China_REE_Refineries.xlsx',
        'ref_urls':     [u for u in (r.get('ref_status'), r.get('ref_capacity'),
                                     r.get('ref_upstream'), r.get('ref_downstream'))
                         if u and u.startswith(('http://', 'https://'))],
        '_upstream_text':   r.get('upstream'),
        '_downstream_text': r.get('downstream'),
    }
    if override.get('note'):
        node['classification_note'] = override['note']
    return node


def build_china_nodes(rows, existing_nodes, cache):
    """Load the China refineries supplement.

    For rows in CHINA_SITE_SPLITS we emit one node per actual refinery site
    (HQ and trading arms omitted). All siblings share a `group_id` so edges
    referencing the aggregate can fan out to every site with `probable: True`.
    """
    name_idx = {}
    for n in existing_nodes:
        if n.get('name') and n.get('lat') is not None:
            name_idx.setdefault(normalize_name(n['name']), n)
    nodes = []
    for r in rows:
        if r['no'] in CHINA_EXCLUDE:
            continue

        # Multi-site row — emit one refinery node per explicit site.
        if r['no'] in CHINA_SITE_SPLITS:
            group_id = f'cn_{r["no"]}'
            for idx, site in enumerate(CHINA_SITE_SPLITS[r['no']], 1):
                node_id   = f'{group_id}_{idx}'
                seed      = node_id
                node_name = f'{r["name"]} — {site["label"]} site'
                node = _base_china_node(r, node_id, name_override=node_name)
                node['group_id']   = group_id
                node['site_label'] = site['label']
                # Geocode the specific site; falls back to China country centroid.
                resolved = resolve_coord([site['geo_query']], 'China', None, seed, cache)
                if resolved:
                    node['lat'], node['lng'], node['precision'] = (
                        round(resolved[0], 5), round(resolved[1], 5), resolved[2])
                    node['geocoded'] = True
                nodes.append(node)
            continue

        # Single-site row — same as before: try upstream inheritance, then
        # Nominatim on the HQ / first-listed location.
        node = _base_china_node(r, f'cn_{r["no"]}')
        seed = node['id']
        lat = lng = precision = None

        up = r.get('upstream')
        if up and node['type'] != 'project':
            primary = re.split(r',| and ', up)[0].strip()
            hit = name_idx.get(normalize_name(primary)) if primary else None
            # China rows are in China — only inherit coords from a Chinese
            # upstream node. Cross-border upstreams (e.g. 'Mountain Pass') must
            # not relocate the plant out of China.
            if hit and hit.get('country') == 'China':
                lat, lng = _jitter((hit['lat'], hit['lng']), 'joined', seed)
                precision = 'joined'

        if lat is None:
            loc = r['location'] or ''
            loc_head = re.split(r'[;\n]', loc)[0]
            loc_head = re.sub(r'^\s*HQ:\s*', '', loc_head, flags=re.I).strip()
            resolved = resolve_coord(
                [loc_head, f'{r["name"]}, China' if r['name'] else None],
                'China', None, seed, cache,
            )
            if resolved:
                lat, lng, precision = resolved

        if lat is not None:
            node.update({'lat': round(lat, 5), 'lng': round(lng, 5),
                         'precision': precision or 'unknown', 'geocoded': True})
        nodes.append(node)
    return nodes


# ── Edges from Factory sheet Upstream/Downstream text ────────────────────────
def build_name_index(all_nodes):
    """Build a {normalized_name: node} index for EXACT-match lookups.

    For multi-site groups (`group_id` set on children), we strip the
    "— <site-label>" tail before indexing so upstream/downstream text that
    names the aggregate ("Southern HREE Smelting & Separation Network") maps
    to a representative child — callers expand to all siblings via
    `expand_to_group`. The first child registered is the representative.
    """
    idx = {}
    for n in all_nodes:
        if not n.get('name'):
            continue
        names = [n['name']]
        if n.get('group_id'):
            group_name = re.sub(r'\s*[—-]\s*.*$', '', n['name'])
            if group_name != n['name']:
                names.append(group_name)
        for nm in names:
            key = normalize_name(nm)
            idx.setdefault(key, n)
    return idx


def resolve_node_ref(text, name_idx):
    """Strict exact-name matcher. No fuzzy scoring, no substring boosts.

    Returns the node whose normalized name equals the normalized input phrase,
    or None. Callers get deterministic, audit-safe edges: if the phrase does
    not literally equal a node name, no edge is emitted.
    """
    if not text:
        return None
    key = normalize_name(text)
    if not key:
        return None
    return name_idx.get(key)


def expand_to_group(node, all_nodes):
    """If `node` belongs to a multi-site group, return every sibling site."""
    gid = node.get('group_id')
    if not gid:
        return [node]
    return [n for n in all_nodes if n.get('group_id') == gid] or [node]


def split_refs(text):
    """Split 'A, B (x, y); C and D' → ['A', 'B', 'C', 'D'] — tolerates nested
    commas inside parens by stripping parens before splitting."""
    if not text:
        return []
    # Drop parenthesized sub-clauses so their commas don't confuse the split.
    stripped = re.sub(r'\s*\([^)]*\)', ' ', str(text))
    parts = re.split(r',|;| and ', stripped)
    return [p.strip() for p in parts if p and p.strip()]


def _fan_out(f_nodes, direction, piece, factory_node, all_nodes, material_hint):
    """Given a matched reference node (or a multi-site group), expand to all
    sites. If the factory itself is multi-site, expand both ends. Emits
    {from, to, probable} tuples suitable for edge building."""
    if not f_nodes:
        return []
    ref_sites = expand_to_group(f_nodes, all_nodes)
    fac_sites = expand_to_group(factory_node, all_nodes)
    # Fan-out count > 1 on either side ⇒ the exact site pairing is uncertain.
    probable = (len(ref_sites) > 1) or (len(fac_sites) > 1)
    out = []
    for r_site in ref_sites:
        for f_site in fac_sites:
            if r_site['id'] == f_site['id']:
                continue
            if direction == 'upstream':
                out.append((r_site, f_site, probable, material_hint(r_site)))
            else:
                out.append((f_site, r_site, probable, material_hint(f_site)))
    return out


def extract_factory_edges(factory_nodes, all_nodes, name_idx):
    def upstream_material(from_node):
        return 'ore' if from_node['id'].startswith('dep_') else 'concentrate'
    def downstream_material(from_node):
        return 'magnet' if from_node.get('type') == 'oem' else 'separated_reo'

    edges = []
    eid = 0
    seen = set()
    for f in factory_nodes:
        gid = f.get('group_id')
        if gid and f['id'] != f'{gid}_1':
            continue
        up = f.get('_upstream_text')
        dn = f.get('_downstream_text')

        for piece in split_refs(up):
            ref = resolve_node_ref(piece, name_idx)
            if not ref:
                continue
            for (a, b, probable, material) in _fan_out(ref, 'upstream', piece, f, all_nodes, upstream_material):
                key = ('up', a['id'], b['id'])
                if key in seen: continue
                seen.add(key)
                eid += 1
                edges.append({
                    'id': f'edge_up_{eid}', 'from_id': a['id'], 'to_id': b['id'],
                    'material': material,
                    'volume_tons_per_year': None, 'year': 2022,
                    'source': 'factory_upstream', 'evidence': piece,
                    'probable': probable,
                })

        for piece in split_refs(dn):
            ref = resolve_node_ref(piece, name_idx)
            if not ref:
                continue
            for (a, b, probable, material) in _fan_out(ref, 'downstream', piece, f, all_nodes, downstream_material):
                key = ('dn', a['id'], b['id'])
                if key in seen: continue
                seen.add(key)
                eid += 1
                edges.append({
                    'id': f'edge_dn_{eid}', 'from_id': a['id'], 'to_id': b['id'],
                    'material': material,
                    'volume_tons_per_year': None, 'year': 2022,
                    'source': 'factory_downstream', 'evidence': piece,
                    'probable': probable,
                })
    return edges


# ── Refinery → Magnet → OEM chain sheets ────────────────────────────────────
# Two small xlsx files in data/raw/ describe aggregate (non-facility-level)
# supply chains: for each row we get a refinery company, a magnet-maker
# company, and one-or-more OEM end-product companies. No lat/lng in the
# source, so we geocode company names via a curated HQ/plant table below.

CHAIN_XLSX = [
    ROOT / 'data' / 'raw' / 'Chinese data from refinery to Companies (Tesla,etc..) BULLSHIT.xlsx',
    ROOT / 'data' / 'raw' / 'Non Chinese data from refinery to Companies (Apple, etc...).xlsx',
]

# Real HQ / primary-plant coordinates for the companies that appear in the
# two chain xlsx files. Single source of truth — swap one entry to relocate
# every node tied to that company. Keep small and curated; don't auto-geocode
# company names (too many false positives: "Lynas" vs. random "Lynas Lake").
COMPANY_HQ = {
    # Chinese magnet makers
    'JL MAG Rare-Earth Co., Ltd.':         (25.8627, 114.9351, 'Ganzhou, Jiangxi, China'),
    'Zhongke Sanhuan High-Tech':           (39.9042, 116.4074, 'Beijing, China'),
    'Ningbo Yunsheng Co., Ltd.':           (29.8683, 121.5440, 'Ningbo, Zhejiang, China'),
    # North American / EU magnet makers
    'MP Materials':                        (35.4778, -115.5311, 'Mountain Pass, California, USA'),  # primary plant
    'E-VAC Magnetics (VAC Group)':         (33.9207,  -80.3414, 'Sumter, South Carolina, USA'),
    'E-VAC Magnetics':                     (33.9207,  -80.3414, 'Sumter, South Carolina, USA'),
    'Noveon Magnetics':                    (29.8833,  -97.9414, 'San Marcos, Texas, USA'),
    'Neo Performance Materials':           (59.3987,   27.7636, 'Sillamäe, Estonia'),
    'HyProMag':                            (52.4862,   -1.8904, 'Birmingham, UK'),
    'USA Rare Earth':                      (36.1156,  -97.0586, 'Stillwater, Oklahoma, USA'),
    # Upstream extras mentioned in the non-Chinese file
    'Energy Fuels':                        (37.5247, -109.4706, 'White Mesa Mill, Utah, USA'),
    'Less Common Metals (LCM)':            (53.2793,   -2.8937, 'Ellesmere Port, UK'),
    'Less Common Metals':                  (53.2793,   -2.8937, 'Ellesmere Port, UK'),
    'VAC Group / E-VAC supply chain':      (50.1355,    8.9147, 'Hanau, Germany'),
    'Caremag/Carester':                    (45.7578,    4.8320, 'Lyon, France'),
    'Caremag':                             (45.7578,    4.8320, 'Lyon, France'),
    'Carester':                            (45.7578,    4.8320, 'Lyon, France'),
    'Lynas Rare Earths':                   (-31.9523,  115.8613, 'Perth, Western Australia'),
    'Australian Strategic Materials':      (-33.8688,  151.2093, 'Sydney, NSW, Australia'),
    'Neo Performance Materials (Silmet)':  (59.3987,   27.7636, 'Sillamäe, Estonia'),
    # Chinese upstream mega-groups (already present as cn_* — link by exact name)
    'China Northern Rare Earth Group':     (40.6186,  109.9405, 'Baotou, Inner Mongolia, China'),
    'China Northern Rare Earth Group (Baotou)': (40.6186, 109.9405, 'Baotou, Inner Mongolia, China'),
    'China Southern Rare Earth Group':     (25.8627,  114.9351, 'Ganzhou, Jiangxi, China'),
    # OEMs
    'Tesla':                               (30.2224,  -97.6197, 'Austin, Texas, USA'),
    'BYD':                                 (22.5431,  114.0579, 'Shenzhen, Guangdong, China'),
    'Toyota':                              (35.0825,  137.1562, 'Toyota City, Aichi, Japan'),
    'General Motors':                      (42.3314,  -83.0458, 'Detroit, Michigan, USA'),
    'Apple':                               (37.3229, -122.0321, 'Cupertino, California, USA'),
    'BMW':                                 (48.1766,   11.5561, 'Munich, Germany'),
    'Jaguar Land Rover':                   (52.4068,   -1.5197, 'Coventry, UK'),
    'Nidec':                               (35.0116,  135.7681, 'Kyoto, Japan'),
    'Nidec Motor Corp':                    (38.6270,  -90.1994, 'St. Louis, Missouri, USA'),
}


def _split_companies(s, extra_seps=False):
    """Split 'Tesla; BYD; Toyota' → ['Tesla','BYD','Toyota']. Strips trailing
    parenthetical asides like 'Tesla (cited customers)'. With extra_seps=True,
    also splits on '+' and ' and ' (useful for the mine column which uses
    'Bayan Obo mine + Baotou separation/alloy facilities')."""
    if not s:
        return []
    pattern = r'[;,]'
    if extra_seps:
        pattern = r'[;,+]| and '
    out = []
    for part in re.split(pattern, s):
        p = part.strip()
        p = re.sub(r'\s*\([^)]*\)\s*$', '', p)  # drop '(cited customers)' tails
        if p and p.lower() not in ('not all disclosed', 'target', 'others'):
            out.append(p)
    return out


def _resolve_company_coords(name):
    """Return (lat, lng, location_text, precision) for a company name, or
    None if we don't have curated coords. No fuzzy matching."""
    hq = COMPANY_HQ.get(name)
    if hq:
        return (hq[0], hq[1], hq[2], 'company_hq')
    # Permit a single canonical alias resolution (e.g. a trailing ", Ltd."
    # stripped). Compare exact after normalization.
    norm = normalize_name(name)
    for canonical, hq in COMPANY_HQ.items():
        if normalize_name(canonical) == norm:
            return (hq[0], hq[1], hq[2], 'company_hq')
    return None


def load_chain_xlsx(path):
    """Load a chain-supply xlsx (refiner/mine/magnet/eem columns). Tolerates
    the two slight schema variants we've seen:
      v1: upstream_company · upstream_asset · upstream_product · magnet_company · oem_company · oem_segment
      v2: refiners         · mines          · refiner product  · magnet_manufacturer · eem_company · eem_segment
    """
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(h) or '' for h in rows[0]]
    # Map both v1 and v2 column names to our canonical keys.
    aliases = {
        'region':           ['region'],
        'mine_name':        ['mines', 'upstream_asset'],
        'refiner_name':     ['refiners', 'upstream_company'],
        'upstream_country': ['upstream_country'],
        'refiner_product':  ['refiner product', 'refiner_product', 'upstream_product'],
        'magnet_company':   ['magnet_manufacturer', 'magnet_company'],
        'magnet_country':   ['magnet_country'],
        'magnet_product':   ['magnet_product'],
        'oem_company':      ['eem_company', 'oem_company'],
        'oem_segment':      ['eem_segment', 'oem_segment'],
        'sources':          ['sources'],
    }
    idx = {}
    for canonical, names in aliases.items():
        for n in names:
            if n in headers:
                idx[canonical] = headers.index(n)
                break
    out = []
    for r in rows[1:]:
        if r is None or r[0] is None:
            continue
        def g(key): return clean(r[idx[key]]) if key in idx else None
        out.append({
            'region':            g('region'),
            'mine_name':         g('mine_name'),
            'refiner_name':      g('refiner_name'),
            'upstream_country':  g('upstream_country'),
            'refiner_product':   g('refiner_product'),
            'magnet_company':    g('magnet_company'),
            'magnet_country':    g('magnet_country'),
            'magnet_product':    g('magnet_product'),
            'oem_company':       g('oem_company'),
            'oem_segment':       g('oem_segment'),
            'sources':           g('sources'),
            'source_file':       path.name,
        })
    return out


def _chain_confidence(source_file):
    """Non-Chinese chain sheet = verified; Chinese chain sheet = inferred.
    Watch out for the 'Non Chinese…' filename also containing 'Chinese data'."""
    sf = source_file or ''
    if sf.startswith('Non '):
        return 'verified'
    if sf.startswith('Chinese'):
        return 'inferred'
    return 'verified'


def build_chain_nodes_and_edges(chain_rows, existing_nodes):
    """Turn each chain row into {upstream? → magnet_maker → oems} + edges.

    Upstream resolution order:
      1. Exact name match against an existing refinery/project (e.g. 'Lynas
         Rare Earths' → the Lynas Kuantan refinery).
      2. COMPANY_HQ lookup, creating a new refinery-type node.
    Magnet maker and OEMs go through COMPANY_HQ only (missing entries are
    skipped so we don't fabricate coords).

    Every edge is flagged `probable: True, source: 'chain_xlsx'` — the source
    files describe aggregate / reputational flows, not facility manifests.
    """
    name_idx = build_name_index(existing_nodes)
    new_nodes = []
    edges = []
    # Dedupe created nodes: same company name across multiple rows → single node.
    by_company_type = {}

    def make_or_reuse(company, node_type, row_source_file, extra=None):
        key = (normalize_name(company), node_type)
        if key in by_company_type:
            return by_company_type[key]
        # (1) reuse an existing node (refinery/project only makes sense upstream).
        if node_type == 'refinery':
            hit = name_idx.get(normalize_name(company))
            if hit and hit['type'] in ('refinery', 'project'):
                by_company_type[key] = hit
                return hit
        hq = _resolve_company_coords(company)
        if not hq:
            return None
        lat, lng, location_text, precision = hq
        prefix = {'refinery': 'chain_ref', 'magnet_maker': 'mag', 'oem': 'oem'}[node_type]
        nid = f'{prefix}_{len([k for k in by_company_type if k[1] == node_type]) + 1}'
        node = {
            'id': nid,
            'type': node_type,
            'name': company,
            'lat': round(lat, 5),
            'lng': round(lng, 5),
            'country': country_from_location(location_text),
            'geocoded': True,
            'precision': precision,
            'confidence': _chain_confidence(row_source_file),
            'company': company,
            'location_text': location_text,
            'source_file': row_source_file or 'chain_xlsx',
            'ref_urls': [],
        }
        if extra:
            node.update(extra)
        new_nodes.append(node)
        by_company_type[key] = node
        return node

    eid = 0
    for r in chain_rows:
        confidence = _chain_confidence(r.get('source_file'))
        sources = r.get('sources')

        # Row expresses a 4-stage chain: mines → refiners → magnet makers → OEMs.
        # Mine strings use '+' between sites (e.g. 'Bayan Obo mine + Baotou…').
        mine_list     = _split_companies(r.get('mine_name'), extra_seps=True)
        refiner_list  = _split_companies(r.get('refiner_name'))
        mag_company   = r.get('magnet_company')
        oem_list      = _split_companies(r.get('oem_company'))

        mag_node = make_or_reuse(mag_company, 'magnet_maker', r.get('source_file'),
                                 extra={'products': r.get('magnet_product')}) if mag_company else None

        # Integrated operators (e.g. MP Materials = refiner + magnet) get a
        # single magnet-maker node; don't duplicate as refiner.
        refiner_nodes = []
        for rc in refiner_list:
            if mag_company and normalize_name(rc) == normalize_name(mag_company):
                continue
            n = make_or_reuse(rc, 'refinery', r.get('source_file'),
                              extra={'products': r.get('refiner_product')})
            if n:
                refiner_nodes.append(n)

        def _strip_suffix(s):
            # Drop generic trailing words so 'Bayan Obo mine' matches 'Bayan Obo'.
            return re.sub(r'\s+(mine|mines|deposit|project|plant|facility|facilities)\s*$', '',
                          s, flags=re.I).strip()

        mine_nodes = []
        for mc in mine_list:
            # Try to reuse an existing mine/deposit/project by exact name first.
            for candidate in (mc, _strip_suffix(mc)):
                hit = name_idx.get(normalize_name(candidate))
                if hit and hit['type'] in ('deposit', 'project'):
                    mine_nodes.append(hit)
                    break
            else:
                hit = None
            if hit:
                continue
            # Otherwise synthesize a project node from COMPANY_HQ (if known).
            n = make_or_reuse(mc, 'refinery', r.get('source_file'))
            # Downgrade synthesized refiner node to 'project' — mines column is
            # about the extraction site, not a processing plant.
            if n and n.get('source_file', '').endswith('chain_xlsx') is False:
                pass  # reused an existing deposit/project; no mutation
            elif n and n['id'].startswith('chain_ref_'):
                n['type'] = 'project'
            if n:
                mine_nodes.append(n)

        oem_nodes = []
        for oc in oem_list:
            n = make_or_reuse(oc, 'oem', r.get('source_file'),
                              extra={'products': r.get('oem_segment')})
            if n:
                oem_nodes.append(n)

        def emit(from_node, to_node, material, direction_tag):
            nonlocal eid
            if not from_node or not to_node or from_node['id'] == to_node['id']:
                return
            eid += 1
            edges.append({
                'id': f'edge_chain_{direction_tag}_{eid}',
                'from_id': from_node['id'], 'to_id': to_node['id'],
                'material': material,
                'volume_tons_per_year': None, 'year': 2024,
                'source': 'chain_xlsx',
                'evidence': f'{r["source_file"]}: {from_node["name"]} → {to_node["name"]}',
                'probable': True,
                'confidence': confidence,
                'sources_text': sources,
            })

        # Mine → refinery (fallback: mine → magnet maker if refiner missing).
        for mn in mine_nodes:
            if refiner_nodes:
                for rn in refiner_nodes:
                    emit(mn, rn, 'ore', 'mr')
            elif mag_node:
                emit(mn, mag_node, 'ore', 'mg')

        # Refinery → magnet maker.
        if mag_node:
            for rn in refiner_nodes:
                emit(rn, mag_node, 'separated_reo', 'rm')

        # Magnet maker → each OEM.
        if mag_node:
            for oem in oem_nodes:
                emit(mag_node, oem, 'magnet', 'dn')
    return new_nodes, edges


def run():
    PROC.mkdir(parents=True, exist_ok=True)
    WEB.mkdir(parents=True, exist_ok=True)
    cache = _load_cache()
    cache_start = len(cache)

    print(f'Loading occurrences …')
    occs = load_occurrences(OCC_XLSX)
    print(f'Loading projects …')
    projects = load_projects(CO_XLSX)
    print(f'Loading factories …')
    factories = load_factories(CO_XLSX)
    print(f'Loading China refineries supplement …')
    china_rows = load_china_refineries(CN_XLSX) if CN_XLSX.exists() else []
    print(f'  occurrences: {len(occs)}')
    print(f'  projects   : {len(projects)}')
    print(f'  factories  : {len(factories)}')
    print(f'  china rows : {len(china_rows)}')

    deposit_nodes = build_deposit_nodes(occs, cache)
    project_nodes, join_edges = build_project_nodes(projects, deposit_nodes, cache)
    factory_nodes = build_factory_nodes(factories, deposit_nodes + project_nodes, cache)
    china_nodes = build_china_nodes(china_rows,
                                    deposit_nodes + project_nodes + factory_nodes,
                                    cache)

    # Apply explicit coordinate overrides (tail of bad source strings). Also
    # re-derive the country from the override location so the UI label is
    # correct (e.g. Kuantan Plant should read "Malaysia", not "Kuantan").
    overrode = 0
    for node in factory_nodes + china_nodes + project_nodes + deposit_nodes:
        o = LOCATION_OVERRIDES.get(node['id'])
        if not o:
            continue
        node['lat'], node['lng'] = o[0], o[1]
        node['precision'] = 'manual_override'
        node['geocoded']  = True
        node['location_text'] = node.get('location_text') or o[2]
        parsed_country = country_from_location(o[2])
        if parsed_country:
            node['country'] = parsed_country
        overrode += 1

    # Default confidence: anything not already tagged is treated as verified.
    # USGS deposits, REE-projects sheet rows, Factory sheet plants, and name-join
    # project↔deposit edges are all sourced from auditable public filings.
    for n in deposit_nodes + project_nodes + factory_nodes:
        n.setdefault('confidence', 'verified')
    for e in join_edges:
        e.setdefault('confidence', 'verified')

    all_nodes = deposit_nodes + project_nodes + factory_nodes + china_nodes
    rendered_nodes = [n for n in all_nodes if n.get('geocoded')]

    name_idx = build_name_index(all_nodes)
    factory_edges = extract_factory_edges(factory_nodes, all_nodes, name_idx)
    for e in factory_edges:
        e['confidence'] = 'verified'  # company upstream/downstream filings
    # Same upstream/downstream extractor works for the China rows since they
    # carry _upstream_text / _downstream_text in the same shape.
    china_edges = extract_factory_edges(china_nodes, all_nodes, name_idx)
    for e in china_edges:
        e['confidence'] = 'inferred'  # China refineries xlsx is user-inferred

    # Refinery → magnet → OEM aggregate chain sheets (supplemental data).
    print('Loading chain xlsx files …')
    chain_rows = []
    for p in CHAIN_XLSX:
        rows = load_chain_xlsx(p)
        print(f'  {p.name}: {len(rows)} rows')
        chain_rows.extend(rows)
    chain_nodes, chain_edges = build_chain_nodes_and_edges(chain_rows, all_nodes)
    all_nodes += chain_nodes
    rendered_nodes = [n for n in all_nodes if n.get('geocoded')]

    edges = join_edges + factory_edges + china_edges + chain_edges

    # Strip internal underscore-prefixed fields before serializing.
    for n in all_nodes:
        for k in list(n):
            if k.startswith('_'):
                n.pop(k)

    meta = {
        'generated_at':   time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()),
        'counts': {
            'deposits':       sum(1 for n in all_nodes if n['type'] == 'deposit'),
            'projects':       sum(1 for n in all_nodes if n['type'] == 'project'),
            'refineries':     sum(1 for n in all_nodes if n['type'] == 'refinery'),
            'magnet_makers':  sum(1 for n in all_nodes if n['type'] == 'magnet_maker'),
            'oem':            sum(1 for n in all_nodes if n['type'] == 'oem'),
            'renderable':     len(rendered_nodes),
            'edges':          len(edges),
            'join_edges':     len(join_edges),
            'factory_edges':  len(factory_edges),
            'china_edges':    len(china_edges),
            'chain_edges':    len(chain_edges),
            'china_rows':     len(china_nodes),
            'chain_nodes':    len(chain_nodes),
        },
        'schema_version': 3,
    }

    (PROC / 'nodes.json').write_text(json.dumps(all_nodes, separators=(',', ':'), ensure_ascii=False))
    (PROC / 'edges.json').write_text(json.dumps(edges,     separators=(',', ':'), ensure_ascii=False))
    bundle = {'meta': meta, 'nodes': all_nodes, 'edges': edges}
    (PROC / 'supply_chain.json').write_text(json.dumps(bundle, separators=(',', ':'), ensure_ascii=False))
    (WEB  / 'supply_chain.json').write_text(json.dumps(bundle, separators=(',', ':'), ensure_ascii=False))

    _save_cache(cache)
    print('\n── Counts ──')
    for k, v in meta['counts'].items():
        print(f'  {k:14s}: {v:>5}')
    print(f'  geocode cache : {len(cache)} entries (+{len(cache) - cache_start})')
    print(f'  manual overrides applied: {overrode}')
    print('\n── Files ──')
    print(f'  {PROC / "nodes.json"}            ({(PROC / "nodes.json").stat().st_size:,} B)')
    print(f'  {PROC / "edges.json"}            ({(PROC / "edges.json").stat().st_size:,} B)')
    print(f'  {PROC / "supply_chain.json"}     ({(PROC / "supply_chain.json").stat().st_size:,} B)')
    print(f'  {WEB  / "supply_chain.json"}     (mirror for the webapp)')


if __name__ == '__main__':
    run()
