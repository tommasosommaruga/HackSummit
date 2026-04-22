"""
Build the processed REE dataset from the raw USGS xlsx.

Inputs:
  data/raw/Global_REE_combined.xlsx
    - Sheet "All Occurrences"  (n≈3,935) — the deposit records
    - Sheet "Abbreviations"    (n≈180)   — token → full-name dictionary
    - Sheet "References"       (n≈1,592) — short_ref → full citation

Outputs:
  data/processed/abbreviations.json     — {token: full_name}
  data/processed/references.json        — {short_ref: full_citation}
  data/processed/occurrences.json       — one record per site, with the
                                          abbreviation and reference joins
                                          materialized as `commods_full` and
                                          `refs_full` arrays
  webapp/public/ree.json                — slim projection used by the webapp
                                          (same fields + joins, 240-char
                                          comment truncation).

Source: USGS Global Rare Earth Element Deposits Database (public domain).
        https://mrdata.usgs.gov/ree/
"""
import json
import math
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW  = ROOT / 'data' / 'raw' / 'Global_REE_combined.xlsx'
PROC = ROOT / 'data' / 'processed'
WEB  = ROOT / 'webapp' / 'public' / 'ree.json'
GEOCACHE = PROC / 'geocode_cache.json'

# ── fields to carry forward into JSON ─────────────────────────────────────────
KEEP = {
    'ID_No':      'id',
    'Name':       'name',
    'Country':    'country',
    'State_Prov': 'state',
    'Latitude':   'lat',
    'Longitude':  'lon',
    'Status':     'status',
    'P_Status':   'pstatus',
    'Rec_Type':   'rtype',
    'Dep_Type':   'dtype',
    'Commods':    'commods',
    'REE':        'ree',
    'REE_Mins':   'ree_mins',
    'Sig_Mins':   'sig_mins',
    'Oth_Mins':   'oth_mins',
    'Host_Lith':  'host_lith',
    'Company':    'company',
    'Region':     'region',
    'Comments':   'comments',
    'P_Years':    'pyears',
    'Discov_Yr':  'discovery',
    'Ref_List':   'refs',
}

# ── centroid tables for points with missing coordinates ───────────────────────
COUNTRY_CENTROIDS = {
    'Algeria': (28.03, 1.66), 'Argentina': (-38.42, -63.62), 'Armenia': (40.07, 45.04),
    'Australia': (-25.27, 133.78), 'Bangladesh': (23.68, 90.36), 'Belarus': (53.71, 27.95),
    'Benin': (9.31, 2.32), 'Bolivia': (-16.29, -63.59), 'Brazil': (-14.24, -51.93),
    'Burundi': (-3.37, 29.92), 'Canada': (56.13, -106.35),
    'Central African Republic': (6.61, 20.94), 'Chile': (-35.68, -71.54), 'China': (35.86, 104.20),
    'Colombia': (4.57, -74.30), 'Czech Republic': (49.82, 15.47),
    'Democratic Republic of Congo': (-4.04, 21.76),
    'Democratic Republic of the Congo': (-4.04, 21.76),
    'Zaire': (-4.04, 21.76), 'Egypt': (26.82, 30.80), 'Ethiopia': (9.14, 40.49),
    'Finland': (61.92, 25.75), 'France': (46.23, 2.21), 'Gabon': (-0.80, 11.61),
    'Germany': (51.17, 10.45), 'Greece': (39.07, 21.82), 'Greenland': (71.71, -42.60),
    'Guyana': (4.86, -58.93), 'India': (20.59, 78.96), 'Indonesia': (-0.79, 113.92),
    'Iran': (32.43, 53.69), 'Italy': (41.87, 12.57), 'Jamaica': (18.11, -77.30),
    'Japan': (36.20, 138.25), 'Kazakhstan': (48.02, 66.92), 'Kenya': (-0.02, 37.91),
    'Kyrgyzstan': (41.20, 74.77), 'Latvia': (56.88, 24.60), 'Madagascar': (-18.77, 46.87),
    'Malawi': (-13.25, 34.30), 'Malaysia': (4.21, 101.98), 'Mexico': (23.63, -102.55),
    'Mongolia': (46.86, 103.85), 'Montenegro': (42.71, 19.37), 'Mozambique': (-18.67, 35.53),
    'Myanmar': (21.91, 95.96), 'Namibia': (-22.96, 18.49), 'Nepal': (28.39, 84.12),
    'New Zealand': (-40.90, 174.89), 'Nigeria': (9.08, 8.68), 'Norway': (60.47, 8.47),
    'Paraguay': (-23.44, -58.44), 'Peru': (-9.19, -75.02), 'Romania': (45.94, 24.97),
    'Russian Federation': (61.52, 105.32), 'Russian Federation(?)': (61.52, 105.32),
    'Rwanda': (-1.94, 29.87),
    'Saint Helena, Ascension, and Tristan da Cunha': (-15.96, -5.71),
    'Saudi Arabia': (23.89, 45.08), 'Sierra Leone': (8.46, -11.78),
    'South Africa': (-30.56, 22.94), 'South Korea': (35.91, 127.77), 'Spain': (40.46, -3.75),
    'Sri Lanka': (7.87, 80.77), 'Sudan': (12.86, 30.22), 'Swaziland': (-26.52, 31.47),
    'Sweden': (60.13, 18.64), 'Taiwan': (23.70, 120.96), 'Tajikistan': (38.86, 71.28),
    'Tanzania': (-6.37, 34.89), 'Thailand': (15.87, 100.99), 'Tibet': (31.69, 88.09),
    'Tunisia': (33.89, 9.54), 'Turkey': (38.96, 35.24), 'Uganda': (1.37, 32.29),
    'Ukraine': (48.38, 31.17), 'United Arab Emirates': (23.42, 53.85),
    'United Kingdom': (55.38, -3.44), 'United States': (37.09, -95.71),
    'Uzbekistan': (41.38, 64.59), 'Venezuela': (6.42, -66.59), 'Vietnam': (14.06, 108.28),
    'Zambia': (-13.13, 27.85), 'Zimbabwe': (-19.02, 29.15),
}

STATE_CENTROIDS = {
    ('China', 'Anhui'): (31.83, 117.23), ('China', 'Beijing'): (39.90, 116.41),
    ('China', 'Chongqing'): (29.43, 106.91), ('China', 'Fujian'): (26.08, 119.30),
    ('China', 'Gansu'): (36.06, 103.83), ('China', 'Guangdong'): (23.13, 113.27),
    ('China', 'Guangxi'): (23.65, 108.32), ('China', 'Guizhou'): (26.58, 106.72),
    ('China', 'Hainan'): (19.19, 109.74), ('China', 'Hebei'): (38.04, 114.50),
    ('China', 'Heilongjiang'): (45.75, 126.64), ('China', 'Henan'): (34.76, 113.66),
    ('China', 'Hubei'): (30.54, 114.34), ('China', 'Hunan'): (28.11, 112.98),
    ('China', 'Inner Mongolia'): (40.82, 111.65), ('China', 'Jiangsu'): (32.06, 118.78),
    ('China', 'Jiangxi'): (28.68, 115.86), ('China', 'Jilin'): (43.90, 125.33),
    ('China', 'Liaoning'): (41.80, 123.43), ('China', 'Ningxia'): (38.47, 106.27),
    ('China', 'Qinghai'): (36.62, 101.78), ('China', 'Shaanxi'): (34.34, 108.94),
    ('China', 'Shandong'): (36.67, 117.02), ('China', 'Shanghai'): (31.23, 121.47),
    ('China', 'Shanxi'): (37.87, 112.55), ('China', 'Sichuan'): (30.57, 104.07),
    ('China', 'Tianjin'): (39.13, 117.20), ('China', 'Tibet'): (29.65, 91.11),
    ('China', 'Xinjiang'): (43.79, 87.62), ('China', 'Yunnan'): (25.04, 102.72),
    ('China', 'Zhejiang'): (30.27, 120.15),
    ('Australia', 'New South Wales'): (-33.87, 151.21), ('Australia', 'Queensland'): (-22.58, 144.07),
    ('Australia', 'South Australia'): (-30.00, 136.21), ('Australia', 'Tasmania'): (-41.45, 145.97),
    ('Australia', 'Victoria'): (-37.47, 144.78), ('Australia', 'Western Australia'): (-27.67, 121.63),
    ('Australia', 'Northern Territory'): (-19.49, 132.55),
    ('Canada', 'Alberta'): (53.93, -116.58), ('Canada', 'British Columbia'): (53.73, -127.65),
    ('Canada', 'Manitoba'): (53.76, -98.81), ('Canada', 'New Brunswick'): (46.57, -66.46),
    ('Canada', 'Newfoundland'): (53.14, -57.66),
    ('Canada', 'Newfoundland and Labrador'): (53.14, -57.66),
    ('Canada', 'Northwest Territories'): (64.83, -124.84),
    ('Canada', 'Nova Scotia'): (44.68, -63.74), ('Canada', 'Nunavut'): (70.30, -83.11),
    ('Canada', 'Ontario'): (51.25, -85.32), ('Canada', 'Prince Edward Island'): (46.51, -63.42),
    ('Canada', 'Quebec'): (52.94, -73.55), ('Canada', 'Saskatchewan'): (52.94, -106.45),
    ('Canada', 'Yukon'): (64.28, -135.00),
    ('Russian Federation', 'Krasnoyarsk'): (64.20, 95.14),
    ('Russian Federation', 'Krasnoyarsk Krai'): (64.20, 95.14),
    ('Russian Federation', 'Murmansk'): (67.90, 33.09),
    ('Russian Federation', 'Murmansk Oblast'): (67.90, 33.09),
    ('Russian Federation', 'Yakutia'): (66.96, 129.66),
    ('Russian Federation', 'Sakha Republic'): (66.96, 129.66),
    ('Russian Federation', 'Irkutsk'): (56.65, 104.33),
    ('Russian Federation', 'Irkutsk Oblast'): (56.65, 104.33),
    ('Russian Federation', 'Karelia'): (62.82, 33.92),
    ('Russian Federation', 'Republic of Karelia'): (62.82, 33.92),
    ('Russian Federation', 'Chukotka'): (66.00, 171.00),
    ('Russian Federation', 'Primorsky Krai'): (45.05, 135.20),
    ('Russian Federation', 'Khabarovsk'): (48.48, 135.08),
    ('Russian Federation', 'Buryatia'): (53.06, 109.13),
    ('India', 'Andhra Pradesh'): (15.91, 79.74), ('India', 'Bihar'): (25.10, 85.31),
    ('India', 'Chhattisgarh'): (21.28, 81.87), ('India', 'Gujarat'): (22.26, 71.19),
    ('India', 'Jharkhand'): (23.61, 85.28), ('India', 'Karnataka'): (15.32, 75.71),
    ('India', 'Kerala'): (10.85, 76.27), ('India', 'Madhya Pradesh'): (22.97, 78.66),
    ('India', 'Maharashtra'): (19.75, 75.71), ('India', 'Odisha'): (20.95, 85.10),
    ('India', 'Orissa'): (20.95, 85.10), ('India', 'Rajasthan'): (27.02, 74.22),
    ('India', 'Tamil Nadu'): (11.13, 78.66), ('India', 'Telangana'): (18.11, 79.02),
    ('India', 'Uttar Pradesh'): (26.85, 80.95), ('India', 'West Bengal'): (22.99, 87.86),
    ('India', 'Meghalaya'): (25.47, 91.37), ('India', 'Assam'): (26.20, 92.94),
    ('Brazil', 'Amazonas'): (-3.42, -65.86), ('Brazil', 'Bahia'): (-12.97, -38.51),
    ('Brazil', 'Ceara'): (-5.50, -39.00), ('Brazil', 'Ceará'): (-5.50, -39.00),
    ('Brazil', 'Goias'): (-15.93, -50.14), ('Brazil', 'Goiás'): (-15.93, -50.14),
    ('Brazil', 'Maranhao'): (-5.00, -45.00), ('Brazil', 'Maranhão'): (-5.00, -45.00),
    ('Brazil', 'Mato Grosso'): (-12.64, -55.42), ('Brazil', 'Minas Gerais'): (-18.51, -44.55),
    ('Brazil', 'Para'): (-3.80, -52.48), ('Brazil', 'Pará'): (-3.80, -52.48),
    ('Brazil', 'Parana'): (-24.89, -51.55), ('Brazil', 'Paraná'): (-24.89, -51.55),
    ('Brazil', 'Rio de Janeiro'): (-22.91, -43.72),
    ('Brazil', 'Rio Grande do Norte'): (-5.79, -36.80),
    ('Brazil', 'Rio Grande do Sul'): (-30.03, -51.22),
    ('Brazil', 'Santa Catarina'): (-27.60, -48.55),
    ('Brazil', 'Sao Paulo'): (-23.55, -46.63), ('Brazil', 'São Paulo'): (-23.55, -46.63),
    ('United States', 'Alaska'): (64.20, -149.49),
    ('United States', 'Arizona'): (34.05, -111.09),
    ('United States', 'California'): (36.78, -119.42),
    ('United States', 'Colorado'): (39.55, -105.78),
    ('United States', 'Idaho'): (44.07, -114.74),
    ('United States', 'Montana'): (46.88, -110.36),
    ('United States', 'Nevada'): (38.80, -116.42),
    ('United States', 'New Mexico'): (34.52, -105.87),
    ('United States', 'North Carolina'): (35.76, -79.02),
    ('United States', 'Texas'): (31.97, -99.90),
    ('United States', 'Utah'): (39.32, -111.09),
    ('United States', 'Washington'): (47.75, -120.74),
    ('United States', 'Wyoming'): (43.08, -107.29),
    ('United States', 'Oregon'): (43.80, -120.55),
}


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s not in ('nan', 'NaN', 'None') else None


# ── Nominatim geocoder (cached) ──────────────────────────────────────────────
# Nominatim policy: <=1 req/sec, set a descriptive User-Agent, and cache.
_LAST_REQ = [0.0]
_UA = 'HackSummit-REE-Geocoder/1.0 (https://github.com/mmsellam)'


def _load_cache():
    if GEOCACHE.exists():
        return json.loads(GEOCACHE.read_text())
    return {}


def _save_cache(cache):
    GEOCACHE.parent.mkdir(parents=True, exist_ok=True)
    GEOCACHE.write_text(json.dumps(cache, indent=2, ensure_ascii=False, sort_keys=True))


def nominatim_lookup(query, cache):
    if query in cache:
        return cache[query]  # may be None (negative cache)
    wait = 1.05 - (time.monotonic() - _LAST_REQ[0])
    if wait > 0:
        time.sleep(wait)
    url = ('https://nominatim.openstreetmap.org/search?'
           + urllib.parse.urlencode({'q': query, 'format': 'json', 'limit': 1}))
    req = urllib.request.Request(url, headers={'User-Agent': _UA})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            results = json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        print(f'  nominatim error for {query!r}: {e}')
        cache[query] = None
        _LAST_REQ[0] = time.monotonic()
        return None
    _LAST_REQ[0] = time.monotonic()
    if not results:
        cache[query] = None
        return None
    hit = results[0]
    try:
        coord = [float(hit['lat']), float(hit['lon'])]
    except (KeyError, ValueError):
        cache[query] = None
        return None
    cache[query] = coord
    return coord


def infer_coord(country, state, id_no, cache):
    """Resolve a best-effort (lat, lon, level) for a point with no USGS coord.

    Priority:
      1. hardcoded STATE_CENTROIDS (stable, zero-network)
      2. Nominatim lookup of "{state}, {country}" (cached)
      3. hardcoded COUNTRY_CENTROIDS
    Level reported to the UI: 'state' (precise within ~1 province) or
    'country' (rough fallback — the whole country centroid).
    """
    base, level = None, None
    if state and country:
        key = (country, state)
        if key in STATE_CENTROIDS:
            base, level = STATE_CENTROIDS[key], 'state'
        else:
            geo = nominatim_lookup(f'{state}, {country}', cache)
            if geo:
                base, level = tuple(geo), 'state'
    if base is None and country in COUNTRY_CENTROIDS:
        base, level = COUNTRY_CENTROIDS[country], 'country'
    if base is None:
        return None
    # Deterministic jitter from id so positions are stable and points don't
    # stack on a single pixel. State hits get tight jitter (~0.25°), country
    # fallbacks get wider jitter (~1.2°) to signal lower confidence visually.
    jitter = 0.25 if level == 'state' else 1.2
    n = 0
    for ch in str(id_no or country):
        n = (n * 131 + ord(ch)) & 0xffffff
    angle = (n * 137.508) * math.pi / 180.0
    r = jitter * math.sqrt(((n >> 8) & 0xff) / 255.0)
    return (base[0] + r * math.sin(angle), base[1] + r * math.cos(angle), level)


# ── sheet loaders / joins ─────────────────────────────────────────────────────
def load_abbreviations(wb):
    ws = wb['Abbreviations']
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header
            continue
        abbr = clean(row[0])
        meaning = clean(row[1])
        if abbr and meaning:
            out[abbr] = meaning
    return out


def load_references(wb):
    ws = wb['References']
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        short = clean(row[0])
        full = clean(row[1])
        if short and full:
            out[short] = full
    return out


def expand_commods(commods_str, abbr):
    """Split 'V, Pb, Zn, Cu, REE' → [('V','vanadium'), …] using the Abbreviations join."""
    if not commods_str:
        return []
    out = []
    for tok in commods_str.split(','):
        t = tok.strip()
        if not t:
            continue
        full = abbr.get(t)  # exact match first
        if not full:
            full = abbr.get(t.upper()) or abbr.get(t.lower())
        out.append({'code': t, 'name': full} if full else {'code': t})
    return out


def resolve_refs(ref_list, refs):
    """Split 'A (2006); B (2010); …' → [{short, full?}, …] using the References join."""
    if not ref_list:
        return []
    out = []
    for piece in ref_list.split(';'):
        s = piece.strip().rstrip(',')
        if not s:
            continue
        full = refs.get(s)
        out.append({'short': s, 'full': full} if full else {'short': s})
    return out


def run():
    if not RAW.exists():
        sys.exit(f'Missing {RAW}. Place the xlsx in data/raw/.')
    PROC.mkdir(parents=True, exist_ok=True)
    WEB.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(RAW, read_only=True, data_only=True)
    abbr = load_abbreviations(wb)
    refs = load_references(wb)
    (PROC / 'abbreviations.json').write_text(
        json.dumps(abbr, indent=2, ensure_ascii=False))
    (PROC / 'references.json').write_text(
        json.dumps(refs, indent=2, ensure_ascii=False))

    ws = wb['All Occurrences']
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    idx = {h: i for i, h in enumerate(headers)}

    cache = _load_cache()
    cache_hits_at_start = len(cache)

    out = []
    stats = {'exact': 0, 'state': 0, 'country': 0, 'dropped': 0}
    for row in rows:
        rec = {}
        for src, dst in KEEP.items():
            v = clean(row[idx[src]])
            if v is not None:
                rec[dst] = v

        # Geocoding: exact from USGS, else inferred from country/state centroid.
        try:
            lat = float(row[idx['Latitude']])
            lon = float(row[idx['Longitude']])
            if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                raise ValueError
            rec['lat'] = round(lat, 4)
            rec['lon'] = round(lon, 4)
            rec['precision'] = 'exact'
            stats['exact'] += 1
        except (TypeError, ValueError):
            inferred = infer_coord(rec.get('country'), rec.get('state'), rec.get('id'), cache)
            if inferred is None:
                stats['dropped'] += 1
                continue
            rec['lat'] = round(inferred[0], 4)
            rec['lon'] = round(inferred[1], 4)
            rec['precision'] = inferred[2]
            stats[inferred[2]] += 1

        # Materialize the Abbreviations + References joins.
        rec['commods_full'] = expand_commods(rec.get('commods'), abbr)
        rec['refs_full']    = resolve_refs(rec.get('refs'), refs)

        out.append(rec)

    # Canonical processed file keeps full comments AND the materialized joins
    # (commods_full, refs_full) so downstream analysis doesn't need to re-join.
    (PROC / 'occurrences.json').write_text(
        json.dumps(out, separators=(',', ':'), ensure_ascii=False))

    # Webapp copy: slim projection. The client fetches abbreviations.json and
    # references.json once and joins at render time, so we strip the inlined
    # full-text arrays to keep the bundle small.
    slim = []
    for r in out:
        s = {k: v for k, v in r.items() if k not in ('commods_full', 'refs_full')}
        if s.get('comments') and len(s['comments']) > 240:
            s['comments'] = s['comments'][:240].rstrip() + '…'
        slim.append(s)
    WEB.write_text(json.dumps(slim, separators=(',', ':'), ensure_ascii=False))

    # Ship the two lookup tables alongside ree.json so the webapp can fetch them.
    (WEB.parent / 'abbreviations.json').write_text(
        json.dumps(abbr, separators=(',', ':'), ensure_ascii=False))
    (WEB.parent / 'references.json').write_text(
        json.dumps(refs, separators=(',', ':'), ensure_ascii=False))

    _save_cache(cache)
    new_lookups = len(cache) - cache_hits_at_start

    print('── Joins built ──')
    print(f'  abbreviations.json : {len(abbr):>5} tokens')
    print(f'  references.json    : {len(refs):>5} citations')
    print(f'  occurrences.json   : {len(out):>5} rows')
    print(f'  geocode cache      : {len(cache):>5} entries (+{new_lookups} new)')
    print('── Geocoding ──')
    for k in ('exact', 'state', 'country', 'dropped'):
        print(f'  {k:<7}: {stats[k]:>5}')
    print('── Files written ──')
    print(f'  {PROC / "abbreviations.json"}')
    print(f'  {PROC / "references.json"}')
    print(f'  {PROC / "occurrences.json"}  ({(PROC / "occurrences.json").stat().st_size:,} bytes)')
    print(f'  {WEB}  ({WEB.stat().st_size:,} bytes)')


if __name__ == '__main__':
    run()
